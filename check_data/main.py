#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
【プレビュー用】元の取得/正規化ロジック(get/normal)をそのまま使い、
指定AND条件を施した「正規化済みデータ」を Excel に最大 5000 行まで出力する。

条件（AND）:
  - gen_col == 1
  - Ess_To_Standard == 1
  - (scope == "jp") の場合: Country_Of_Registration == "JP JAPAN"

出力:
  - 6シート（all/jp × 3g/4g/5g）
  - 各シート最大 5000 行
  - ヘッダ固定（1行目をFreeze）
  - フォント: メイリオ 10pt
  - ヘッダは「CSVの全ヘッダ」（内部列 __src_rownum などは除外）

注意:
- 正規化内容（改行削除、COMP_LEGAL_NAME の ',' '.' 除去等）は normalization.py 側の normal() に依存します。
  （＝normalization.py を修正済み前提）

依存:
  pip install openpyxl
"""

from __future__ import annotations

from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


import sys
import os
# このファイルのパスを基準に、../std をパスに追加
current_dir = os.path.dirname(os.path.abspath(__file__))
std_path = os.path.abspath(os.path.join(current_dir, '..', 'std'))
sys.path.insert(0, std_path)

from table_sql import get, apply_pipeline, qident
from table_rule import Pipeline
from normalization import normal



# ============================================================
# 設定
# ============================================================
CSV_PATH = "../../ISLD-export/ISLD-export.csv"
DB_PATH = "work.sqlite"

OUT_XLSX = "preview_norm_filtered.xlsx"
MAX_ROWS_PER_SHEET = 5000

JP_VALUE = "JP JAPAN"
SCOPE_ALL = "all"
SCOPE_JP = "jp"

GEN_COLS = ("3G", "4G", "5G")
GEN_KEY = {"3G": "3g", "4G": "4g", "5G": "5g"}


# ============================================================
# ヘルパ
# ============================================================
def is_jp_scope(scope: str) -> bool:
    return scope == SCOPE_JP


def pick_output_headers(all_cols: List[str]) -> List[str]:
    """
    Excelに出したい「CSV本来のヘッダ」を返す。
    - 内部列（例: __src_rownum）は除外
    """
    return [c for c in all_cols if not c.startswith("__")]


def build_filtered_table(raw_norm, *, scope: str, gen_col: str, out_table: str, return_heads: List[str]):
    """
    条件（AND）:
      - gen_col == 1
      - Ess_To_Standard == 1
      - (scope == "jp") の場合: Country_Of_Registration == "JP JAPAN"
    """
    p = Pipeline()
    p.where_eq(gen_col, 1)
    p.where_eq("Ess_To_Standard", 1)
    if is_jp_scope(scope):
        p.where_eq("Country_Of_Registration", JP_VALUE)

    return apply_pipeline(raw_norm, p, out_table_name=out_table, return_heads=return_heads)


def fetch_rows_for_excel(tbl, *, export_heads: List[str], limit: int) -> List[List[str]]:
    """
    テーブルから先頭 limit 行を取得して、Excel用に list[list[str]] にする。
    - __src_rownum があれば __src_rownum で並べて取り出す（元CSV順を保つ）
    """
    cols = tbl.columns()
    has_src = "__src_rownum" in cols

    # SELECT列は「__src_rownum + export_heads」（ただしExcelには export_heads だけ出す）
    select_heads = (["__src_rownum"] if has_src else []) + export_heads

    sel_cols = ", ".join(qident(c) for c in select_heads)
    from_tbl = qident(tbl.table_name)

    if has_src:
        sql = f"SELECT {sel_cols} FROM {from_tbl} ORDER BY __src_rownum LIMIT {int(limit)}"
    else:
        sql = f"SELECT {sel_cols} FROM {from_tbl} LIMIT {int(limit)}"

    out: List[List[str]] = []
    for row in tbl.iterquery(sql):
        # row: (__src_rownum?, col1, col2, ...)
        if has_src:
            row = row[1:]
        out.append(["" if v is None else str(v) for v in row])
    return out


def autosize_columns(ws, *, max_width: int = 60, min_width: int = 10, sample_rows: int = 200) -> None:
    """重くしない程度に列幅を推定する（先頭 sample_rows 行のみ）。"""
    max_row = min(ws.max_row, sample_rows)
    max_col = ws.max_column

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_width, max(min_width, max_len + 2))


def write_excel(
    out_path: str,
    *,
    headers: List[str],
    rows_by_sheet: Dict[str, List[List[str]]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    font_body = Font(name="メイリオ", size=10)
    font_head = Font(name="メイリオ", size=10, bold=True)
    align = Alignment(vertical="top", wrap_text=True)

    for sheet_name, rows in rows_by_sheet.items():
        ws = wb.create_sheet(title=sheet_name)

        # ヘッダ
        ws.append(headers)
        for cell in ws[1]:
            cell.font = font_head
            cell.alignment = align

        # データ
        for r in rows:
            ws.append(r)

        # 本文フォント
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = font_body
                cell.alignment = align

        # ヘッダ固定
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        autosize_columns(ws)

    wb.save(out_path)


# ============================================================
# main
# ============================================================
def main() -> None:
    # 1) 元の取得（get）
    raw = get(CSV_PATH, db_path=DB_PATH)  # head指定しない＝全列取得（ヘッダ全取得要件）
    try:
        # 2) 元の正規化（normal）※ normalization.py の定義を使用
        raw_norm = normal(raw, out_table_name="t_norm")

        # 3) ヘッダ（CSV全ヘッダ：内部列除外）
        norm_cols = raw_norm.columns()
        export_heads = pick_output_headers(norm_cols)

        # 4) 6通りを抽出してExcel化
        targets: List[Tuple[str, str]] = [
            (SCOPE_ALL, "3G"),
            (SCOPE_ALL, "4G"),
            (SCOPE_ALL, "5G"),
            (SCOPE_JP,  "3G"),
            (SCOPE_JP,  "4G"),
            (SCOPE_JP,  "5G"),
        ]

        rows_by_sheet: Dict[str, List[List[str]]] = {}

        for scope, gen_col in targets:
            gen = GEN_KEY[gen_col]
            sheet = f"{scope}_{gen}"

            # フィルタ済みテーブルを作成（返す列は __src_rownum + export_heads にして順序維持可能に）
            return_heads = (["__src_rownum"] if "__src_rownum" in norm_cols else []) + export_heads
            t_flt = build_filtered_table(
                raw_norm,
                scope=scope,
                gen_col=gen_col,
                out_table=f"t_preview_{scope}_{gen}",
                return_heads=return_heads,
            )

            # 先頭MAX_ROWS_PER_SHEETだけ取得して、Excel用に整形
            rows_by_sheet[sheet] = fetch_rows_for_excel(t_flt, export_heads=export_heads, limit=MAX_ROWS_PER_SHEET)

        # 5) Excel 書き出し（メイリオ10pt + ヘッダ固定）
        write_excel(OUT_XLSX, headers=export_heads, rows_by_sheet=rows_by_sheet)
        print(f"OK: wrote {OUT_XLSX}")

    finally:
        raw.close()


if __name__ == "__main__":
    main()