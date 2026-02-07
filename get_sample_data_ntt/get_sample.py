#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import csv
import sys
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook

# ----------------------------------------------------------------------
# ローカルライブラリパス（../std）を追加（指定どおり）
# ----------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
LIB_DIR = (SCRIPT_DIR / ".." / "std").resolve()
sys.path.insert(0, str(LIB_DIR))

from table_sql import get, build_pipeline_sql  # noqa: E402
from table_rule import Pipeline                # noqa: E402
from progress import Progress                  # noqa: E402
from normalization import normal               # noqa: E402


# ============================================================
# ★指定はここにまとめる（全部 Optional）
#   - None にするとその抽出はスキップ
# ============================================================
TARGET_DIPG_ID: Optional[int] = 43483
TARGET_ETPR_ID: Optional[int] = 876
TARGET_DIPG_PATF_ID: Optional[int] = 20438
TARGET_IPRD_REFERENCE: Optional[str] = "ISLD-201608-010"

# サンプル行数（サンプル系ジョブにだけ適用）
N = 100

# 会社固定（このスクリプトの主役）
COMP_COL = "COMP_LEGAL_NAME"
COMP_VAL = "NTT DOCOMO INC"
COMPANY_TAG = "NTT"

# ============================================================
# 固定入出力
# ============================================================
SOURCE_CSV = (SCRIPT_DIR / "../../ISLD-export/ISLD-export.csv").resolve()
SQLITE_DB = (SCRIPT_DIR / "work.sqlite").resolve()


# ★保存先を ../<COMP_VAL>/ の中にする（Windowsでも安全な名前にする）
def _safe_dir_name(s: str) -> str:
    bad = '<>:"/\\|?*'
    s2 = "".join("_" if c in bad else c for c in str(s))
    s2 = " ".join(s2.split()).strip()  # 連続空白整理
    return s2 or "COMP"


OUT_DIR = (SCRIPT_DIR / ".." / _safe_dir_name(COMP_VAL)).resolve()
OUT_DIR.mkdir(parents=True, exist_ok=True)


# ユーザー指定：この列すべてを読み＆出力
SOURCE_COLUMNS = [
    "IPRD_ID",
    "IPRD_REFERENCE",
    "IPRD_SIGNATURE_DATE",
    "Reflected_Date",
    "COMP_LEGAL_NAME",
    "DECL_IS_PROP_FLAG",
    "LICD_DEC_PREP_TO_GRANT_FLAG",
    "LICD_REC_CONDI_FLAG",
    "DIPG_ID",
    "DIPG_DISPLAY_NUMBER",
    "DIPG_EXTERNAL_ID",
    "Standard",
    "Ess_To_Standard",
    "WI_Type",
    "WOIT_ETSI_DELIVERABLE_NUMBER",
    "WOIT_REFERENCE",
    "WOIT_VERSION",
    "3GPP_Type",
    "TGPP_NUMBER",
    "TGPV_VERSION",
    "Patent_Type",
    "PATT_APPLICATION_NUMBER",
    "PUBL_NUMBER",
    "PBPA_TITLEEN",
    "PBPA_PRIORITY_NUMBERS",
    "Country_Of_Registration",
    "PBPA_APP_DATE",
    "ETPR_ID",
    "ETPR_ACRONYM",
    "Ess_To_Project",
    "Illustrative_Part",
    "DIPG_PATF_ID",
    "Original_Application_Number",
    "Original_Publication_Number",
    "Explicitely_Disclosed",
    "Normalized_Patent",
    "2G",
    "3G",
    "4G",
    "5G",
]


# ============================================================
# Progress 互換（API差分に耐える）
# ============================================================
def _progress_new(total: int, every_lines: int) -> Optional[Any]:
    try:
        return Progress(file_size_bytes=total, progress_every_lines=every_lines)
    except Exception:
        pass
    try:
        return Progress(file_size=total, every=every_lines)
    except Exception:
        pass
    try:
        return Progress(file_size_bytes=total)
    except Exception:
        pass
    try:
        return Progress(file_size=total)
    except Exception:
        return None


def _progress_tick(prog: Any, *, lines: int, table_name: str = "") -> None:
    if prog is None:
        return
    fn = getattr(prog, "tick", None)
    if callable(fn):
        try:
            fn(lines_total=lines, table_name=table_name)
            return
        except TypeError:
            pass
        except Exception:
            return
    fn2 = getattr(prog, "tick_bytes", None)
    if callable(fn2):
        try:
            fn2(lines)
        except Exception:
            return


def _progress_done(prog: Any, *, lines: int, table_name: str = "") -> None:
    if prog is None:
        return
    fn = getattr(prog, "done", None)
    if callable(fn):
        try:
            fn(lines_total=lines, table_name=table_name)
            return
        except TypeError:
            pass
        except Exception:
            return
        try:
            fn()
        except Exception:
            return


# ============================================================
# Excelシート名（<=31、禁止文字除去、重複回避）
# ============================================================
_INVALID_SHEET_CHARS = set(r'[]:*?/\\')

def _safe_sheet_title(title: str) -> str:
    s = "".join("_" if ch in _INVALID_SHEET_CHARS else ch for ch in title).strip()
    return s or "sheet"

def _make_unique_sheet_title(desired: str, used: set[str]) -> str:
    base = _safe_sheet_title(desired)[:31]
    if base not in used:
        used.add(base)
        return base
    n = 1
    while True:
        suffix = f"_{n}"
        cut = 31 - len(suffix)
        cand = (base[:cut] + suffix)[:31]
        if cand not in used:
            used.add(cand)
            return cand
        n += 1


# ============================================================
# Pipeline を SQL にコンパイルして直接ストリーム取得
# ============================================================
def _iter_rows_by_pipeline(norm_tbl, pipeline: Pipeline, *, limit: Optional[int]):
    plan = build_pipeline_sql(pipeline, return_heads=SOURCE_COLUMNS)
    sql = plan.sql.format(table=f'"{norm_tbl.table_name}"') + " ORDER BY __src_rownum ASC"
    params = list(plan.params)
    if limit is not None:
        sql += " LIMIT ?"
        params.append(limit)
    # 返り値は (__src_rownum, *SOURCE_COLUMNS)
    return norm_tbl.conn.execute(sql, tuple(params))


def _write_csv_from_cursor(out_csv: Path, cursor, *, show_progress: bool, progress_total: Optional[int] = None) -> int:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    total = progress_total if progress_total is not None else N
    prog = _progress_new(total=total, every_lines=10) if show_progress else None

    written = 0
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(SOURCE_COLUMNS)
        for row in cursor:
            w.writerow(list(row[1:]))  # row[0] は __src_rownum
            written += 1
            _progress_tick(prog, lines=written, table_name=out_csv.name)

    _progress_done(prog, lines=written, table_name=out_csv.name)
    return written


def _csv_to_excel_sheet(ws, csv_path: Path, *, max_rows: int = 1_048_576) -> int:
    rows = 0
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.reader(f)
        for rec in r:
            ws.append(rec)
            rows += 1
            if rows >= max_rows:
                break
    return rows


# ============================================================
# main
# ============================================================
def main() -> None:
    if not SOURCE_CSV.exists():
        raise FileNotFoundError(f"not found: {SOURCE_CSV}")

    # --------------------------
    # ジョブ定義
    # --------------------------
    jobs: list[tuple[str, Pipeline, Optional[int]]] = []

    def add_job(tag: str, pipe: Pipeline, limit: Optional[int]) -> None:
        jobs.append((tag, pipe, limit))

    # 01: 会社サンプル（★where_eq(COMP_COL, COMP_VAL) は絶対に残す）
    add_job(
        f"{COMPANY_TAG}_{N}",
        Pipeline().where_eq(COMP_COL, COMP_VAL),
        N,
    )

    # 02-: 「キー=値 の全件抽出」系（似てるのでまとめる）
    targets_all = [
        ("TARGET_IPRD_REFERENCE", "IPRD_REFERENCE", TARGET_IPRD_REFERENCE),
        ("TARGET_DIPG_ID", "DIPG_ID", TARGET_DIPG_ID),
        ("TARGET_ETPR_ID", "ETPR_ID", TARGET_ETPR_ID),
        ("TARGET_DIPG_PATF_ID", "DIPG_PATF_ID", TARGET_DIPG_PATF_ID),
    ]
    for tag, col, val in targets_all:
        if val is not None:
            add_job(
                f"{tag}__{val}__ALL",
                Pipeline()
                    .where_eq(COMP_COL, COMP_VAL)
                    .where_eq(col, val),
                None,
            )

    # 条件付きサンプル（会社条件は必ず入れる）
    add_job(
        f"{COMPANY_TAG}_JP_{N}",
        Pipeline()
            .where_eq(COMP_COL, COMP_VAL)
            .where_eq("Country_Of_Registration", "JP JAPAN"),
        N,
    )

    add_job(
        f"{COMPANY_TAG}_JP_UNIQ_PBPA_PRIORITY_NUMBERS_{N}",
        Pipeline()
            .where_eq(COMP_COL, COMP_VAL)
            .where_eq("Country_Of_Registration", "JP JAPAN")
            .unique_by("PBPA_PRIORITY_NUMBERS"),
        N,
    )

    add_job(
        f"{COMPANY_TAG}_JP_UNIQ_PATT_APPLICATION_NUMBER_{N}",
        Pipeline()
            .where_eq(COMP_COL, COMP_VAL)
            .where_eq("Country_Of_Registration", "JP JAPAN")
            .unique_by("PATT_APPLICATION_NUMBER"),
        N,
    )

    add_job(
        f"{COMPANY_TAG}_BASIS_PATENT_{N}",
        Pipeline().where_eq(COMP_COL, COMP_VAL).where_eq("Patent_Type", "Basis Patent"),
        N,
    )

    # --------------------------
    # get + normal は1回だけ
    # --------------------------
    raw = get(str(SOURCE_CSV), db_path=str(SQLITE_DB), head=SOURCE_COLUMNS)
    try:
        norm_tbl = normal(raw, out_table_name="t_norm")

        out_csv_paths: list[Path] = []

        for i, (tag, pipe, lim) in enumerate(jobs, start=1):
            csv_name = f"{i:02d}_{tag}.csv"
            out_csv = (OUT_DIR / csv_name).resolve()

            cur = _iter_rows_by_pipeline(norm_tbl, pipe, limit=lim)

            show_prog = (lim is not None)  # サンプル（N指定）のときだけプログレス表示
            written = _write_csv_from_cursor(
                out_csv,
                cur,
                show_progress=show_prog,
                progress_total=lim,
            )

            out_csv_paths.append(out_csv)
            print(f"[OK] {out_csv.name}: wrote {written} rows")

        # Excel作成（各CSV名＝シート名）
        xlsx_path = (OUT_DIR / f"samples_{COMPANY_TAG}.xlsx").resolve()
        wb = Workbook(write_only=True)

        used_titles: set[str] = set()
        for p in out_csv_paths:
            desired_title = p.stem
            title = _make_unique_sheet_title(desired_title, used_titles)
            ws = wb.create_sheet(title=title)
            _csv_to_excel_sheet(ws, p)

        wb.save(xlsx_path)
        print(f"[OK] Excel saved: {xlsx_path}")

    finally:
        raw.close()


if __name__ == "__main__":
    main()
