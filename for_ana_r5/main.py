#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import sys
import csv
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook

# ----------------------------------------------------------------------
# ローカルライブラリパス（../std）を追加
# ----------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
LIB_DIR = (SCRIPT_DIR / ".." / "std").resolve()
sys.path.insert(0, str(LIB_DIR))

from table_sql import get, apply_pipeline
from table_rule import Pipeline
from normalization import normal
from table_al import TableAL


# ============================================================
# 入出力設定
# ============================================================
SOURCE_CSV = Path("../../ISLD-export/ISLD-export.csv")
SQLITE_DB = Path("work.sqlite")

OUT_DIR = Path("out_isld_extract_and_fd")
OUT_DIR.mkdir(parents=True, exist_ok=True)

OUT_EXTRACT = OUT_DIR / "extract"
OUT_FD = OUT_DIR / "fd_global"
OUT_EXTRACT.mkdir(parents=True, exist_ok=True)
OUT_FD.mkdir(parents=True, exist_ok=True)

GEN_FLAGS = ("3G", "4G", "5G")
GEN_SUFFIX = {"3G": "3g", "4G": "4g", "5G": "5g"}

SCOPE_ALL = "all"
SCOPE_JP = "jp"
JP_COUNTRY_VALUE = "JP JAPAN"

# 会社FDで使う unique キー
UNIQ_KEYS = ("IPRD_REFERENCE", "IPRD_ID", "DIPG_ID", "DIPG_PATF_ID")

# ============================================================
# SOURCE_COLUMNS（あなた指定）
# ============================================================
SOURCE_COLUMNS = [
    "IPRD_ID",
    "IPRD_REFERENCE",
    "IPRD_SIGNATURE_DATE",
    "Reflected_Date",
    "COMP_LEGAL_NAME",
    "DECL_IS_PROP_FLAG",
    "LICD_DEC_PREP_TO_GRANT_FLAG",
    "LICD_REC_CONDI_FLAG",
    "DIPG_ID",
    "DIPG_DISPLAY_NUMBER",
    "DIPG_EXTERNAL_ID",
    "Standard",
    "Ess_To_Standard",
    "WI_Type",
    "WOIT_ETSI_DELIVERABLE_NUMBER",
    "WOIT_REFERENCE",
    "WOIT_VERSION",
    "3GPP_Type",
    "TGPP_NUMBER",
    "TGPV_VERSION",
    "Patent_Type",
    "PATT_APPLICATION_NUMBER",
    "PUBL_NUMBER",
    "PBPA_TITLEEN",
    "PBPA_PRIORITY_NUMBERS",
    "Country_Of_Registration",
    "PBPA_APP_DATE",
    "ETPR_ID",
    "ETPR_ACRONYM",
    "Ess_To_Project",
    "Illustrative_Part",
    "DIPG_PATF_ID",
    "Original_Application_Number",
    "Original_Publication_Number",
    "Explicitely_Disclosed",
    "Normalized_Patent",
    "2G",
    "3G",
    "4G",
    "5G",
]


# ============================================================
# SQLite 識別子クォート & テーブルアクセス
# ============================================================
def _table_name(tbl) -> str:
    for a in ("table_name", "name", "_table_name", "_name"):
        if hasattr(tbl, a):
            v = getattr(tbl, a)
            if isinstance(v, str) and v:
                return v
    raise RuntimeError("Cannot detect table name from table object.")


def _conn(tbl):
    for a in ("conn", "_conn", "connection"):
        if hasattr(tbl, a):
            c = getattr(tbl, a)
            if hasattr(c, "execute"):
                return c
    if hasattr(tbl, "db"):
        db = getattr(tbl, "db")
        for a in ("conn", "_conn", "connection"):
            if hasattr(db, a):
                c = getattr(db, a)
                if hasattr(c, "execute"):
                    return c
    raise RuntimeError("Cannot detect sqlite connection from table object.")


def _qident(name: str) -> str:
    # SQLite identifier quote
    return '"' + str(name).replace('"', '""') + '"'


def _iter_table_rows(tbl, requested_columns: list[str]):
    """
    SQLite上のtblから、存在する列だけ抽出して
    (cols, cursor) を返す（列名と行イテレータ）
    """
    tname = _table_name(tbl)
    conn = _conn(tbl)
    qt = _qident(tname)

    cur = conn.execute(f"PRAGMA table_info({qt})")
    existing = {row[1] for row in cur.fetchall()}  # row[1]=colname

    cols = [c for c in requested_columns if c in existing]
    if not cols:
        raise RuntimeError(f"No requested columns exist in table: {tname}")

    col_sql = ", ".join(_qident(c) for c in cols)
    sql = f"SELECT {col_sql} FROM {qt}"
    cursor = conn.execute(sql)
    return cols, cursor


# ============================================================
# Excel: シート名生成（31文字制限、禁則文字、重複）
# ============================================================
_INVALID_SHEET_CHARS = set(r'[]:*?/\\')

def _sanitize_sheet_title(title: str) -> str:
    s = "".join("_" if ch in _INVALID_SHEET_CHARS else ch for ch in title)
    s = s.strip()
    return s or "sheet"

def _make_unique_sheet_title(desired: str, used: set[str]) -> str:
    """
    Excel制限:
      - 31文字以内
      - 禁則文字除去
      - 重複回避（末尾に _1, _2 ...）
    """
    base = _sanitize_sheet_title(desired)
    base = base[:31]

    if base not in used:
        used.add(base)
        return base

    # 重複した場合、末尾に _n を付ける（31文字以内）
    n = 1
    while True:
        suffix = f"_{n}"
        cut = 31 - len(suffix)
        cand = (base[:cut] + suffix)[:31]
        if cand not in used:
            used.add(cand)
            return cand
        n += 1


# ============================================================
# CSV出力
# ============================================================
def export_table_to_csv(tbl, out_path: Path, columns: list[str]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # tbl.to_csv があればまず試す
    if hasattr(tbl, "to_csv"):
        try:
            tbl.to_csv(str(out_path))
            return
        except Exception:
            pass

    cols, cursor = _iter_table_rows(tbl, columns)

    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(cols)
        for row in cursor:
            w.writerow(list(row))


# ============================================================
# Excel出力（テーブル→シート）
# ============================================================
def write_table_to_workbook(tbl, wb: Workbook, sheet_csv_name: str, columns: list[str], *, used_titles: set[str]) -> None:
    """
    SQLiteテーブルの内容をWorkbookの1シートへ書き込み（ストリーム）
    """
    title = _make_unique_sheet_title(sheet_csv_name, used_titles)
    ws = wb.create_sheet(title=title)

    cols, cursor = _iter_table_rows(tbl, columns)
    ws.append(cols)
    for row in cursor:
        ws.append(list(row))


def write_fd_to_workbook(wb: Workbook, sheet_csv_name: str, labels: list[Any], counts: list[Any], *, used_titles: set[str]) -> None:
    """
    FD (labels, counts) をWorkbookの1シートへ
    """
    title = _make_unique_sheet_title(sheet_csv_name, used_titles)
    ws = wb.create_sheet(title=title)
    ws.append(["COMP_LEGAL_NAME", "count"])
    for lab, cnt in zip(labels or [], counts or []):
        ws.append([lab, cnt])


# ============================================================
# (Ａ) 抽出（CSV 3本 + まとめExcel 1本）
# ============================================================
def run_extract_csvs(norm_tbl, *, wb_extract: Workbook | None = None) -> None:
    targets = [
        ("ref_ISLD-201608-010", "IPRD_REFERENCE", "ISLD-201608-010"),
        ("dipg_43483",          "DIPG_ID",        43483),
        ("patf_20438",          "DIPG_PATF_ID",   20438),
    ]

    used_titles: set[str] = set()

    for tag, col, val in targets:
        p = Pipeline().where_eq(col, val)
        t = apply_pipeline(norm_tbl, p, out_table_name=f"t_extract__{tag}", return_heads=SOURCE_COLUMNS)

        csv_name = f"extract__{tag}.csv"
        export_table_to_csv(t, OUT_EXTRACT / csv_name, SOURCE_COLUMNS)

        if wb_extract is not None:
            # シート名はCSV名（ただしExcel制限で短縮/重複回避あり）
            write_table_to_workbook(t, wb_extract, csv_name, SOURCE_COLUMNS, used_titles=used_titles)


# ============================================================
# (Ｂ) 会社FD（CSV 24本 + genごとExcel 3本）
# ============================================================
def is_jp(scope: str) -> bool:
    return scope == SCOPE_JP


def make_base_filtered_table(norm_tbl, *, scope: str, gen_flag: str, out_table: str, return_cols: list[str]):
    """
    全体データに共通フィルタだけ適用（抽出条件は入れない）
    """
    p = Pipeline()
    p.where_eq(gen_flag, 1)
    p.where_eq("Ess_To_Standard", 1)
    p.where_eq("Ess_To_Project", 1)
    p.where_ne("DIPG_PATF_ID", -1)
    p.where_eq("Normalized_Patent", 1)
    if is_jp(scope):
        p.where_eq("Country_Of_Registration", JP_COUNTRY_VALUE)

    return apply_pipeline(norm_tbl, p, out_table_name=out_table, return_heads=return_cols)


def make_unique_by_key(tbl, *, uniq_key: str, out_table: str, return_cols: list[str]):
    try:
        tbl.create_index(uniq_key)
    except Exception:
        pass
    p = Pipeline().unique_by(uniq_key)
    return apply_pipeline(tbl, p, out_table_name=out_table, return_heads=return_cols)


def fd_filename(scope: str, gen_suffix: str, uniq_key: str) -> str:
    return f"fd_cmp__{scope}__{gen_suffix}__uniq-by_{uniq_key}.csv"


def run_global_company_fd(norm_tbl, *, al: TableAL, wb_by_gen: dict[str, Workbook] | None = None) -> None:
    cols_needed = ["COMP_LEGAL_NAME"] + list(UNIQ_KEYS) + [
        "Ess_To_Standard", "Ess_To_Project", "DIPG_PATF_ID", "Normalized_Patent",
        "Country_Of_Registration", "3G", "4G", "5G"
    ]

    # Excel sheet title の重複管理（世代ごと）
    used_titles_by_gen: dict[str, set[str]] = {g: set() for g in GEN_FLAGS}

    for scope in (SCOPE_ALL, SCOPE_JP):
        for gen_flag in GEN_FLAGS:
            gen_suffix = GEN_SUFFIX[gen_flag]

            # 1) 全体データに共通フィルタだけ
            t_filtered = make_base_filtered_table(
                norm_tbl,
                scope=scope,
                gen_flag=gen_flag,
                out_table=f"t_global__{scope}__flt__{gen_suffix}",
                return_cols=cols_needed,
            )

            # 2) uniqueキーごとに unique → FD
            for uniq_key in UNIQ_KEYS:
                t_uq = make_unique_by_key(
                    t_filtered,
                    uniq_key=uniq_key,
                    out_table=f"t_global__{scope}__uq_{uniq_key}__{gen_suffix}",
                    return_cols=["COMP_LEGAL_NAME", uniq_key],
                )
                labels, counts = al.frequency_distribution(t_uq, "COMP_LEGAL_NAME")

                csv_name = fd_filename(scope, gen_suffix, uniq_key)
                out_csv = OUT_FD / csv_name
                al.save_as_file((labels, counts), str(out_csv))

                # Excel（genごと）にも書き込み
                if wb_by_gen is not None:
                    wb = wb_by_gen.get(gen_flag)
                    if wb is not None:
                        write_fd_to_workbook(
                            wb,
                            csv_name,
                            labels,
                            counts,
                            used_titles=used_titles_by_gen[gen_flag],
                        )


# ============================================================
# main
# ============================================================
def main() -> None:
    al = TableAL()

    raw = get(str(SOURCE_CSV), db_path=str(SQLITE_DB), head=SOURCE_COLUMNS)
    try:
        norm_tbl = normal(raw, out_table_name="t_norm")

        # ==========================
        # (A) 抽出CSV + Excel(1本)
        # ==========================
        wb_extract = Workbook(write_only=True)
        run_extract_csvs(norm_tbl, wb_extract=wb_extract)
        out_xlsx_a = OUT_EXTRACT / "extract_all.xlsx"
        wb_extract.save(out_xlsx_a)

        # ==========================
        # (B) 会社FD CSV + Excel(gen別3本)
        # ==========================
        wb_by_gen = {g: Workbook(write_only=True) for g in GEN_FLAGS}
        run_global_company_fd(norm_tbl, al=al, wb_by_gen=wb_by_gen)

        for gen_flag in GEN_FLAGS:
            gen_suffix = GEN_SUFFIX[gen_flag]
            out_xlsx_b = OUT_FD / f"fd_global__{gen_suffix}.xlsx"
            wb_by_gen[gen_flag].save(out_xlsx_b)

    finally:
        raw.close()


if __name__ == "__main__":
    main()